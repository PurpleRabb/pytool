#!/usr/bin/env python
# -*- coding:utf-8 -*-

import codecs
import getopt
import glob
import os
import re
import shutil
import subprocess
import sys

from html.entities import entitydefs
from importlib import reload
from urllib.request import urlopen
from html.parser import HTMLParser
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment


def usage():
    print('''################Useage:################
./read_bulletin.py [<options>] <SECURITY_BULLETIN_DATE>\n
SECURITY_BULLETIN_DATE:       The date of the security bulletin
Options:
-g, --google:          Get the google patch
-w, --website          Get google patch from website, default is from locale
-k, --kernel:          Get the kernel website patch
-q, --qcom:            Get the qcom website patch
-e, --excel:           Save excel file.
-h, --help:            Print help message
#######################################''')


class HtmlNoteInfo:
    def __init__(self):
        self.source_link_type = ''
        self.cve_dir_name = ''
        self.html_link = ''
        self.html_ref = ''
        self.aosp_version = ''

    def set_source_link_type(self, source_link_type):
        self.source_link_type = source_link_type

    def set_html_link(self, html_link):
        self.html_link = html_link

    def set_html_ref(self, html_ref):
        self.html_ref = html_ref

    def set_cve_dir_name(self, cve_dir_name):
        self.cve_dir_name = cve_dir_name

    def set_aosp_version(self, aosp_version):
        self.aosp_version = aosp_version


class PatchHtmlParser(HTMLParser):
    def __init__(self, is_patch_file, file_write):
        HTMLParser.__init__(self)
        # Html link address parse
        self.start_a_node = False
        # CVE node parse
        self.start_cve_node = False
        # Need to add new node
        self.new_node = False

        # Patch diff comment node parse
        self.start_comment_class = False
        # Patch file node parse
        self.start_file_diff_class = False
        # Content node parse
        self.start_content_diff_class = False

        # Node a href value
        self.hrefValue = ''
        # All note info list
        self.note_info_data = []
        self.cve_dir_value = ''
        self.is_patch_file = is_patch_file
        self.file_write = file_write
        self.get_google_type = False
        self.get_kernel_type = False
        self.get_qcom_type = False
        if get_patch_type.find('google') >= 0:
            self.get_google_type = True
        if get_patch_type.find('kernel') >= 0:
            self.get_kernel_type = True
        if get_patch_type.find('qcom') >= 0:
            self.get_qcom_type = True

    def handle_starttag(self, tag, attrs):
        print(tag)
        if tag == 'a':
            # print(attrs)
            for k, v in attrs:
                if k == 'href':
                    self.hrefValue = v
                    # print('The v:' + v)

                    if 'android.googlesource.com/platform/' in v:
                        new_html_node = HtmlNoteInfo()
                        new_html_node.set_source_link_type('GOOGLE')
                        self.new_html_node.append(new_html_node)
                        if self.get_google_type:
                            print("########get_google_type")
                            self.start_a_node = True
                    elif 'git.kernel.org' in v:
                        new_html_node = HtmlNoteInfo()
                        new_html_node.set_source_link_type('KERNEL')
                        self.new_html_node.append(new_html_node)
                        if self.get_kernel_type:
                            self.start_a_node = True
                    elif 'source.codeaurora.org' in v:
                        new_html_node = HtmlNoteInfo()
                        new_html_node.set_source_link_type('QCOM')
                        self.new_html_node.append(new_html_node)
                        if self.get_qcom_type:
                            self.start_a_node = True
        elif tag == 'td':
            self.start_cve_node = True
        elif tag == 'tr':
            self.new_html_node = []
        elif tag == 'pre':
            for k, v in attrs:
                if k == 'class' and v == 'u-pre u-monospace MetadataMessage':
                    self.start_comment_class = True
                elif k == 'class' and v == 'u-pre u-monospace Diff':
                    self.start_file_diff_class = True
                elif k == 'class' and v == 'u-pre u-monospace Diff-unified':
                    self.start_content_diff_class = True

    def handle_endtag(self, tag):
        if tag == 'a':
            self.start_a_node = False
        elif tag == 'td':
            self.start_cve_node = False
        elif tag == 'tr':
            if self.new_node:
                for new_html_node in self.new_html_node:
                    # print 'new html node : ' + new_html_node.cve_dir_name + ', ' + new_html_node.aosp_version
                    self.note_info_data.append(new_html_node)
                    self.new_node = False
        elif tag == 'pre':
            if self.start_comment_class:
                self.file_write.write('\n')
            self.start_comment_class = False
            self.start_file_diff_class = False
            self.start_content_diff_class = False

    def handle_data(self, data):
        # print 'The data:' + data
        if self.start_a_node:
            if not self.hrefValue.startswith('http'):
                link = 'https:' + self.hrefValue
            else:
                link = self.hrefValue
            self.new_html_node[-1].set_html_link(link)
            self.new_html_node[-1].set_html_ref(data.strip())
            self.new_html_node[-1].set_cve_dir_name(self.cve_dir_value)
            self.new_node = True
            # self.note_info_data.append(self.new_html_node)
        elif self.start_cve_node:
            cveMatchObj = re.match(r'CVE-', data, re.M | re.I)
            if cveMatchObj:
                self.cve_dir_value = data.strip()
                print('The CVE dir:' + data)
            elif self.new_node:
                aospVersionMatchObj = re.match(r'(\d(\.\d)*,*)+', data)
                if aospVersionMatchObj:
                    if self.new_html_node[0].source_link_type == 'GOOGLE':
                        for new_html_node in self.new_html_node:
                            new_html_node.set_aosp_version(data)
                            # self.note_info_data[-1].set_aosp_version(data)
                            print('AOSP version: ' + data)
        elif self.start_comment_class or self.start_file_diff_class or self.start_content_diff_class:
            self.file_write.write(data)
        elif self.is_patch_file:
            self.file_write.write(data)

    def handle_entityref(self, name):
        if self.start_comment_class or self.start_file_diff_class or self.start_content_diff_class:
            if entitydefs.has_key(name):
                details = entitydefs[name]
            else:
                details = '&%s;' % (name)
            self.file_write.write(details)

    def handle_charref(self, name):
        try:
            charnum = int(name)
        except ValueError:
            return
        if charnum < 1 or charnum > 255:
            return
        if self.start_comment_class or self.start_file_diff_class or self.start_content_diff_class:
            self.file_write.write(chr(charnum))
        print('The charref:' + name)


def html_node_parse(diff_addr, cve_dir_name, file_name, is_patch_file):
    con = urlopen(diff_addr, timeout=100)
    doc = con.read()
    doc.replace(r'a\n', 'a ')
    con.close()
    diff_path = os.path.join(BASE_DIR, cve_dir_name + '\\' + file_name)
    print('The diff path:' + diff_path)
    file_write = codecs.open(diff_path, 'w', 'utf-8')
    parse = PatchHtmlParser(is_patch_file, file_write)
    parse.feed(str(doc))
    file_write.close()
    parse.close()


EXCELL_ROW = 1
bd = Side(style='thin', color='FF000000')
BORDER = Border(left=bd, top=bd, right=bd, bottom=bd)
HIGHLIGHT = Font(underline='single', color='0563C1')


def write_excel(ws, i, html, file_name):
    if get_patch_type.find('excel') >= 0:
        global EXCELL_ROW, BORDER, HIGHLIGHT
        ws.cell(row=EXCELL_ROW, column=1).value = html.cve_dir_name
        # ws.cell(row=EXCELL_ROW, column=2).value = '=HYPERLINK("{}", "{}")'.format(html.html_link, html.html_ref)
        ws.cell(row=EXCELL_ROW, column=2).hyperlink = html.html_link
        ws.cell(row=EXCELL_ROW, column=2).value = html.html_ref
        # ws.cell(row=EXCELL_ROW, column=2).style = HIGHLIGHT
        ws.cell(row=EXCELL_ROW, column=2).font = HIGHLIGHT
        ws.cell(row=EXCELL_ROW, column=3).value = file_name
        ws.cell(row=EXCELL_ROW, column=3).alignment = Alignment(wrap_text=True)  # 自动换行
        ws.cell(row=EXCELL_ROW, column=4).value = html.aosp_version
        ws.cell(row=EXCELL_ROW, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=EXCELL_ROW, column=5).alignment = Alignment(wrap_text=True)
        for col in range(1, 6):
            ws.cell(row=EXCELL_ROW, column=col).border = BORDER
        EXCELL_ROW += 1


def get_bulletin():
    # Get all patch html list
    base_url = 'https://source.android.google.cn/security/bulletin/%s' % (bulletin_date)
    base_url2 = 'https://source.android.com/security/bulletin/%s' % (bulletin_date)
    con = urlopen(base_url2, timeout=100)
    doc = con.read()
    con.close()
    parse = PatchHtmlParser(False, None)
    # print(doc)
    parse.feed(str(doc).replace(r'a\n', 'a '))
    parse.close()

    if get_patch_type.find('excel') >= 0:
        print('excel start')
        if not os.path.exists(BASE_DIR):
            os.makedirs(BASE_DIR)

        wb = Workbook()
        ws = wb.active

        # ws.page_setup.fitToWidth = 1
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30

    i = 0
    for html in parse.note_info_data:
        diff_addr = ''
        file_name = ''
        i += 1
        is_patch_file = False
        cve_dir = os.path.join(BASE_DIR, html.cve_dir_name)
        if not os.path.exists(cve_dir):
            os.makedirs(cve_dir)
        if html.source_link_type == 'GOOGLE':
            diff_addr = html.html_link + '%5E%21/'
            start_pos = html.html_link.find('platform')
            end_pos = 0
            if start_pos >= 0:
                end_pos = html.html_link.find('+')
            # print 'The android start pos: %s, end_pos: %s' % (start_pos, end_pos)
            file_name = html.html_link[start_pos: end_pos]
            if make_goole_locale_patch:
                # remove platform suffix
                source_format_dir = file_name.replace('/', '\\')[9:]
                source_format_dir = os.path.join(GOOGLE_LOCALE_SOURCE_DIR, source_format_dir)
                commit_id = html.html_link[end_pos + 2:]
                print('The base: source_format_dir: %s, the commit id:%s' % (source_format_dir, commit_id))
                # To the source code path
                patch_or_dir_no_exist = False
                if os.path.exists(source_format_dir):
                    os.chdir(source_format_dir)
                    subprocess.Popen('git clean -fxd .', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    cmd = 'git format-patch %s~1..%s' % (commit_id, commit_id)
                    print(cmd)
                    format_patch = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    # os.popen(cmd)
                    # format_patch = subprocess.Popen('git format-patch -1 ' + commit_id, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    format_patch.wait()
                    if format_patch.returncode != 0:
                        patch_or_dir_no_exist = True
                else:
                    patch_or_dir_no_exist = True
                if patch_or_dir_no_exist:
                    print('No patch found in locale for commit id:%s' % (commit_id))
                    file_name = file_name.replace('/', '-')
                    file_name = file_name + 'patch_%s.diff' % (i)
                    html_node_parse(diff_addr, html.cve_dir_name, file_name, is_patch_file)
                    write_excel(ws, i, html, file_name)
                else:
                    print('Find the patch in locale')
                    patch_dir = os.path.join(BASE_DIR, html.cve_dir_name)
                    patch_files = glob.glob('*.patch')
                    patch_files.sort()
                    print('len : ' + str(len(patch_files)))

                    j = 0
                    file_name = file_name.replace('/', '-')
                    if len(patch_files) == 1:
                        patch_file_name = patch_files[j]
                        patch_name = file_name + 'patch_%s.patch' % (i)
                        os.rename(patch_file_name, patch_name)
                        print('patch_dir: %s, file_name: %s' % (patch_dir, patch_name))
                        shutil.move(patch_name, patch_dir)
                        write_excel(ws, i, html, patch_name)
                    else:
                        for patch in patch_files:
                            patch_file_name = patch_files[j]
                            j += 1
                            patch_name = file_name + 'patch_%s_%s.patch' % (i, j)
                            os.rename(patch_file_name, patch_name)
                            print('patch_dir %s: %s, file_name: %s' % (j, patch_dir, patch_name))
                            shutil.move(patch_name, patch_dir)
                            write_excel(ws, i, html, patch_name)

                # To the CWD dir ??
                # os.chdir(CWD_DIR)
                continue
            file_name = file_name.replace('/', '-')
            file_name = file_name + 'patch_%s.diff' % (i)
        elif html.source_link_type == 'KERNEL':
            diff_addr = html.html_link
            diff_addr = diff_addr.replace('commit', 'patch')
            file_name = 'kernel-patch_%s.patch' % (i)
            is_patch_file = True
        elif html.source_link_type == 'QCOM':
            diff_addr = html.html_link
            diff_addr = diff_addr.replace('commit', 'patch')
            start_pos = html.html_link.find('quic/')
            end_pos = 0
            if start_pos >= 0:
                # 5 is the length of 'quic/'
                start_pos = start_pos + 5
                end_pos = html.html_link.find('commit')
            # print 'The qcom start pos: %s, end_pos: %s' % (start_pos, end_pos)
            file_name = html.html_link[start_pos: end_pos]
            file_name = file_name.replace('/', '-')
            file_name = file_name + 'patch_%s.patch' % (i)
            is_patch_file = True
        print('The html addr:' + diff_addr)
        print('file_name:' + file_name)
        html_node_parse(diff_addr, html.cve_dir_name, file_name, is_patch_file)
        write_excel(ws, i, html, file_name)

    if get_patch_type.find('excel') >= 0:
        print('save excel file')
        wb.save(BASE_DIR + '\\' + bulletin_date + '.xlsx')


if __name__ == '__main__':
    try:
        opts, args = getopt.getopt(sys.argv[1:], 'gwkqeh',
                                   ['google=', 'website=', 'kernel=', 'qcom=', 'excel=', 'help='])
    except getopt.GetoptError as err:
        usage()
        exit(1)

    # Please modify to your real directory path
    GOOGLE_LOCALE_SOURCE_DIR = r'X:\r33'
    make_goole_locale_patch = True

    if len(opts) == 0:
        get_patch_type = 'google,kernel,qcom,excel'
    else:
        get_patch_type = ''
    for op, value in opts:
        if op in ('-g', '--google'):
            get_patch_type = 'google'
        elif op in ('-w', '--website'):
            make_goole_locale_patch = False
        elif op in ('-k', '--kernel'):
            get_patch_type += ',kernel'
        elif op in ('-q', '--qcom'):
            get_patch_type += ',qcom'
        elif op in ('-e', '--excel'):
            get_patch_type += 'excel'
        elif op in ('-h', '--help'):
            usage()
            exit(1)
    print('###The patch type:' + get_patch_type)

    if not len(args) == 1:
        bulletin_date = input('Please input security bulletin date:')
        if bulletin_date == '':
            print('Invalid parameters input\n')
            usage()
            os.system("pause")
            exit(1)
    else:
        bulletin_date = args[0].strip()
    print('The bulletin date:' + bulletin_date)

    reload(sys)
    # sys.setdefaultencoding('utf-8')

    CWD_DIR = os.getcwd()
    BASE_DIR = os.path.join(CWD_DIR, bulletin_date)
    if os.path.exists(BASE_DIR):
        shutil.rmtree(BASE_DIR)
    get_bulletin()
